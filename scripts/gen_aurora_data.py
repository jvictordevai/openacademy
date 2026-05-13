"""Generate Aurora Distribuidora sample data: vendas + financeiro xlsx.

Reproducible (fixed seed). Plants anomalies described in the schema:
- Concentration: Vipar Engenharia = ~19% of revenue
- Outlier: Marcos Tavares with -42% ticket vs team
- Margin compression: LED category from 28% → 16% over 8 months
- AR aging deterioration: average delay 11d → 26d
- Cost increase: Prysmian +23% in 6 months
- Freight ratio creep: 2.1% → 3.8% of net revenue
- Cohort retention drop: 71% → 54%
"""

from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

random.seed(42)

OUT_DIR = Path(__file__).resolve().parent.parent / "public" / "materiais"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ============== EMPRESA ==============

COMPANY = {
    "nome": "Distribuidora Aurora Ltda.",
    "cnpj": "29.484.207/0001-65",
    "endereco": "Rod. Raposo Tavares, km 28 — Cotia/SP",
    "filial": "Av. Albino J. Barbosa de Oliveira, 1500 — Campinas/SP",
    "ano_fundacao": 2018,
}

# ============== DIMENSÕES ==============

VENDEDORES = [
    {"id": "V001", "nome": "Carla Mendes",        "regiao": "SP-Capital",  "tipo": "Externo", "data_admissao": "2019-04-22", "meta_mensal": 380000, "comissao_pct": 2.5},
    {"id": "V002", "nome": "Marcos Tavares",      "regiao": "SP-Capital",  "tipo": "Externo", "data_admissao": "2022-01-10", "meta_mensal": 320000, "comissao_pct": 2.5},
    {"id": "V003", "nome": "Patricia Yoshida",    "regiao": "SP-Capital",  "tipo": "Externo", "data_admissao": "2020-09-03", "meta_mensal": 350000, "comissao_pct": 2.5},
    {"id": "V004", "nome": "Roberto Almeida",     "regiao": "SP-Capital",  "tipo": "Externo", "data_admissao": "2021-06-18", "meta_mensal": 340000, "comissao_pct": 2.5},
    {"id": "V005", "nome": "João Bittencourt",    "regiao": "Interior-SP", "tipo": "Externo", "data_admissao": "2019-11-12", "meta_mensal": 280000, "comissao_pct": 2.8},
    {"id": "V006", "nome": "Eduarda Carvalho",    "regiao": "Interior-SP", "tipo": "Externo", "data_admissao": "2023-02-27", "meta_mensal": 260000, "comissao_pct": 2.8},
    {"id": "V007", "nome": "Diego Salgueiro",     "regiao": "Sul-MG",      "tipo": "Externo", "data_admissao": "2022-08-15", "meta_mensal": 240000, "comissao_pct": 3.0},
    {"id": "V008", "nome": "Luana Prado",         "regiao": "Sul-MG",      "tipo": "Externo", "data_admissao": "2024-03-04", "meta_mensal": 200000, "comissao_pct": 3.0},
    {"id": "V009", "nome": "Fernanda Lima",       "regiao": "Inside",      "tipo": "Inside",  "data_admissao": "2021-02-01", "meta_mensal": 180000, "comissao_pct": 1.8},
    {"id": "V010", "nome": "Bruno Sato",          "regiao": "Inside",      "tipo": "Inside",  "data_admissao": "2022-05-09", "meta_mensal": 170000, "comissao_pct": 1.8},
    {"id": "V011", "nome": "Renato Falcão",       "regiao": "Inside",      "tipo": "Inside",  "data_admissao": "2023-10-21", "meta_mensal": 150000, "comissao_pct": 1.8},
]

SEGMENTOS_CLIENTE = [
    ("Construtora", 0.18, (250000, 900000)),
    ("Instaladora", 0.32, (40000, 180000)),
    ("Revenda", 0.28, (60000, 280000)),
    ("Indústria", 0.12, (120000, 500000)),
    ("Setor Público", 0.05, (300000, 800000)),
    ("Outros", 0.05, (20000, 80000)),
]

CIDADES_UF = [
    ("São Paulo", "SP"), ("São Paulo", "SP"), ("São Paulo", "SP"), ("Cotia", "SP"),
    ("Guarulhos", "SP"), ("Osasco", "SP"), ("Santo André", "SP"),
    ("Campinas", "SP"), ("Sorocaba", "SP"), ("Ribeirão Preto", "SP"), ("São José dos Campos", "SP"),
    ("Bauru", "SP"), ("Jundiaí", "SP"),
    ("Belo Horizonte", "MG"), ("Uberlândia", "MG"), ("Juiz de Fora", "MG"),
    ("Curitiba", "PR"), ("Florianópolis", "SC"), ("Porto Alegre", "RS"),
]

CATEGORIAS = {
    "Cabos e Fios": {
        "margem_alvo": 0.32,
        "items": [
            ("Cabo Flex 2,5mm² 100m Prysmian", "RL", 198.40, 289.00, "Prysmian"),
            ("Cabo Flex 4,0mm² 100m Prysmian", "RL", 312.50, 459.00, "Prysmian"),
            ("Cabo Flex 6,0mm² 100m Prysmian", "RL", 488.20, 712.00, "Prysmian"),
            ("Cabo PP 3x1,5mm² 100m Cobrecom", "RL", 142.80, 209.90, "Cobrecom"),
            ("Cabo PP 3x2,5mm² 100m Cobrecom", "RL", 218.40, 318.00, "Cobrecom"),
            ("Cabo Coaxial RG6 100m Furukawa", "RL", 178.00, 264.00, "Furukawa"),
            ("Fio Sólido 1,5mm² 100m Prysmian", "RL", 88.20, 132.90, "Prysmian"),
            ("Fio Sólido 2,5mm² 100m Prysmian", "RL", 142.00, 211.00, "Prysmian"),
            ("Cabo de Rede CAT6 305m Furukawa", "CX", 612.00, 879.00, "Furukawa"),
            ("Cabo Solar 6mm² 100m Solfio", "RL", 387.60, 549.00, "Solfio"),
        ],
    },
    "Iluminação LED": {
        "margem_alvo": 0.28,  # vai cair com tempo
        "items": [
            ("Painel LED 60x60 48W 6500K", "UN", 87.20, 119.90, "Importado-CN"),
            ("Painel LED 30x60 24W 6500K", "UN", 52.40, 74.90, "Importado-CN"),
            ("Lâmpada LED Bulbo 9W E27 Empalux", "UN", 7.80, 14.90, "Empalux"),
            ("Lâmpada LED Bulbo 12W E27 Empalux", "UN", 9.40, 18.90, "Empalux"),
            ("Refletor LED 50W IP65 Avant", "UN", 64.20, 98.00, "Avant"),
            ("Refletor LED 100W IP65 Avant", "UN", 112.40, 169.00, "Avant"),
            ("Spot LED 5W Embutir Empalux", "UN", 8.60, 14.50, "Empalux"),
            ("Fita LED 5m IP65 12V Stella", "UN", 38.20, 64.90, "Stella"),
            ("Luminária Pendente LED 24W", "UN", 78.00, 134.00, "Importado-CN"),
            ("Plafon LED 24W 6500K", "UN", 42.80, 72.00, "Importado-CN"),
        ],
    },
    "Quadros e Disjuntores": {
        "margem_alvo": 0.30,
        "items": [
            ("Quadro Distribuição 24DIN Steck", "UN", 134.50, 198.00, "Steck"),
            ("Quadro Distribuição 12DIN Steck", "UN", 88.20, 132.00, "Steck"),
            ("Disjuntor Monopolar 10A Curva C Steck", "UN", 11.20, 18.90, "Steck"),
            ("Disjuntor Monopolar 25A Curva C Steck", "UN", 11.80, 19.90, "Steck"),
            ("Disjuntor Bipolar 40A Curva C Steck", "UN", 38.40, 62.00, "Steck"),
            ("Disjuntor Tripolar 63A Curva C Schneider", "UN", 124.00, 189.00, "Schneider"),
            ("DR Bipolar 30mA 25A Schneider", "UN", 142.20, 219.00, "Schneider"),
            ("DR Tetrapolar 30mA 40A Schneider", "UN", 287.00, 432.00, "Schneider"),
            ("DPS Classe II 275V 20kA Clamper", "UN", 28.40, 49.00, "Clamper"),
            ("Contator Tripolar 25A WEG", "UN", 184.20, 269.00, "WEG"),
        ],
    },
    "Automação e Infra": {
        "margem_alvo": 0.34,
        "items": [
            ("Eletroduto Flexível 3/4\" 50m Tigre", "RL", 78.40, 119.00, "Tigre"),
            ("Eletroduto Rígido 25mm 3m Tigre", "UN", 24.80, 39.00, "Tigre"),
            ("Caixa de Passagem 200x200x100 PVC", "UN", 48.20, 78.00, "Tramontina"),
            ("Eletrocalha Lisa 100x50x3000 Cemar", "UN", 88.40, 134.00, "Cemar"),
            ("Tomada 2P+T 20A Branca Tramontina", "UN", 12.20, 19.90, "Tramontina"),
            ("Interruptor Simples Tramontina Linha Lux", "UN", 9.80, 16.50, "Tramontina"),
            ("Conector Wago 3 Vias 32A", "UN", 4.80, 8.90, "Wago"),
            ("Sensor de Presença Bivolt Intelbras", "UN", 78.00, 124.00, "Intelbras"),
            ("Câmera IP Bullet 2MP Intelbras", "UN", 312.00, 489.00, "Intelbras"),
            ("Switch 8 portas Gigabit Intelbras", "UN", 218.00, 339.00, "Intelbras"),
        ],
    },
}

NOMES_EMPRESAS = [
    "Vipar Engenharia", "Construbase", "Sólida Construtora", "Cetel Engenharia",
    "Elétrica Marília", "Eletro Sul Materiais", "Mafer Engenharia", "Plaza Empreendimentos",
    "Instaladora Foco", "Norte Iluminação", "Brilho Materiais Elétricos",
    "Concept Engenharia", "Construsul", "TecnoEletro",
    "FastInstale", "Tronon Eletro", "GEL Engenharia", "Energ Tech",
    "Materiais Bonsucesso", "Casa do Eletricista", "Casa das Lâmpadas",
    "Eletroshow", "PrimeWatt", "VoltMax", "Eletrocentral",
    "Mendes Construtora", "Saber Engenharia", "Cinco Estrelas Engenharia",
    "GoodTech Automação", "InfraOne", "Real Iluminação", "Lumen Comercial",
    "Brasil Lumens", "Iluminar SP", "GreenLight Soluções",
    "Andrade & Cia", "VKR Engenharia", "ProEletric",
    "Loja do Eletricista (Campinas)", "Eletrosul Cotia", "Casa Forte Materiais",
    "InfraMaster", "TopWatt Distribuidora",
]

PRIMEIROS_NOMES = ["Marcos", "Ana", "Roberto", "Patrícia", "Eduardo", "Camila", "Fernando", "Juliana", "Ricardo", "Vanessa"]
SOBRENOMES = ["Silva", "Santos", "Oliveira", "Costa", "Pereira", "Almeida", "Rodrigues", "Fernandes"]

# ============== HELPERS ==============

def gen_cnpj(seed_idx: int) -> str:
    """Generate plausible CNPJ format (not real validation)."""
    base = str((10000000 + seed_idx * 71) % 99999999).zfill(8)
    return f"{base[:2]}.{base[2:5]}.{base[5:8]}/0001-{(seed_idx * 13 + 7) % 90 + 10:02d}"


def days_iter(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


def round_money(v: float) -> float:
    return round(v, 2)


def autosize(ws) -> None:
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            try:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
            except Exception:
                continue
        ws.column_dimensions[col_letter].width = min(max_len + 2, 48)


def write_headers(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="0F1828")
    header_font = Font(bold=True, color="FFFFFF", name="Inter")
    thin = Side(border_style="thin", color="2A3447")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = border
    ws.freeze_panes = "A2"


def add_table(ws, ref_last_col: int, last_row: int, name: str) -> None:
    last_letter = get_column_letter(ref_last_col)
    table_ref = f"A1:{last_letter}{last_row}"
    tbl = Table(displayName=name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

# ============== GERAR CLIENTES ==============

def gen_clientes() -> list[dict]:
    clientes = []
    n_target = 380

    # Vipar Engenharia — cliente top concentrador
    clientes.append({
        "cliente_id": "AUR-0001",
        "razao_social": "Vipar Engenharia Ltda",
        "cnpj": "12.345.678/0001-90",
        "segmento": "Construtora",
        "cidade": "São Paulo",
        "uf": "SP",
        "vendedor_responsavel": "Carla Mendes",
        "data_primeira_compra": "2021-03-14",
        "limite_credito": 850000,
        "status": "Ativo",
    })

    for i in range(2, n_target + 1):
        seg_choices = []
        for nome, peso, _ in SEGMENTOS_CLIENTE:
            seg_choices.extend([nome] * int(peso * 100))
        segmento = random.choice(seg_choices)
        seg_data = next((s for s in SEGMENTOS_CLIENTE if s[0] == segmento), SEGMENTOS_CLIENTE[0])
        limite = random.randint(*seg_data[2])

        # construir nome de empresa plausível
        if random.random() < 0.6 and i <= len(NOMES_EMPRESAS) + 1:
            base = NOMES_EMPRESAS[(i - 2) % len(NOMES_EMPRESAS)]
            sufixos = [" Ltda", " ME", " EIRELI", " S/A", " & Cia"]
            razao = base + (sufixos[i % len(sufixos)] if "Ltda" not in base and "ME" not in base else "")
        else:
            primeiro = random.choice(PRIMEIROS_NOMES)
            sobrenome = random.choice(SOBRENOMES)
            tipos = ["Construções", "Engenharia", "Materiais Elétricos", "Instaladora", "Comércio Elétrico"]
            razao = f"{primeiro} {sobrenome} {random.choice(tipos)} Ltda"

        cidade_uf = random.choice(CIDADES_UF)
        vendedor = random.choices(
            [v for v in VENDEDORES if v["tipo"] == "Externo"],
            weights=[3, 2, 3, 3, 3, 2, 2, 1],
        )[0]

        # Marcos Tavares concentra mais clientes pequenos
        if vendedor["nome"] == "Marcos Tavares" and segmento in ("Construtora", "Indústria"):
            if random.random() < 0.7:
                vendedor = random.choice([v for v in VENDEDORES if v["nome"] != "Marcos Tavares" and v["tipo"] == "Externo"])

        # data primeira compra: distribuído 2019 → 2025
        days_back = random.randint(60, 365 * 6)
        first_date = date(2025, 12, 1) - timedelta(days=days_back)

        # ~12% inativos
        status = "Inativo" if random.random() < 0.12 else "Ativo"

        clientes.append({
            "cliente_id": f"AUR-{i:04d}",
            "razao_social": razao,
            "cnpj": gen_cnpj(i),
            "segmento": segmento,
            "cidade": cidade_uf[0],
            "uf": cidade_uf[1],
            "vendedor_responsavel": vendedor["nome"],
            "data_primeira_compra": first_date.isoformat(),
            "limite_credito": limite,
            "status": status,
        })
    return clientes

# ============== GERAR PRODUTOS ==============

def gen_produtos() -> list[dict]:
    produtos = []
    sku_idx = 1
    for cat_nome, cat_data in CATEGORIAS.items():
        for item in cat_data["items"]:
            descricao, unidade, custo, preco, fornecedor = item
            cat_prefix = {
                "Cabos e Fios": "CABO",
                "Iluminação LED": "LED",
                "Quadros e Disjuntores": "QDR",
                "Automação e Infra": "INFRA",
            }[cat_nome]
            sku = f"{cat_prefix}-{sku_idx:04d}"
            sku_idx += 1
            produtos.append({
                "sku": sku,
                "descricao": descricao,
                "categoria": cat_nome,
                "subcategoria": "Geral",
                "unidade": unidade,
                "custo_medio": custo,
                "preco_tabela": preco,
                "margem_alvo_pct": cat_data["margem_alvo"] * 100,
                "fornecedor_principal": fornecedor,
            })
    return produtos

# ============== GERAR VENDAS (24 meses) ==============

def gen_vendas(clientes: list[dict], produtos: list[dict]) -> list[dict]:
    """24 meses de pedidos com sazonalidade + anomalias plantadas."""
    vendas = []
    pedido_id_seq = 100000

    start = date(2024, 1, 1)
    end = date(2025, 12, 31)

    ativos = [c for c in clientes if c["status"] == "Ativo"]
    vipar = clientes[0]  # AUR-0001

    # Distribuir clientes pra vendedores como base
    vendedor_clientes = {v["nome"]: [] for v in VENDEDORES if v["tipo"] == "Externo"}
    for c in ativos:
        if c["vendedor_responsavel"] in vendedor_clientes:
            vendedor_clientes[c["vendedor_responsavel"]].append(c)

    for d in days_iter(start, end):
        # sazonalidade — pico mar/abr (mês 3-4) e set/out (9-10), vale jan/dez
        month = d.month
        season_mult = 1.0
        if month in (3, 4, 9, 10):
            season_mult = 1.35
        elif month in (1, 12):
            season_mult = 0.72
        elif month in (7,):  # férias
            season_mult = 0.82

        # tendência leve de crescimento 2024 → 2025
        year_mult = 1.0 + (0.07 if d.year == 2025 else 0)

        # final de semana — menos vendas
        if d.weekday() >= 5:
            if random.random() > 0.15:
                continue
            season_mult *= 0.4

        # número de pedidos do dia
        n_pedidos = int(random.gauss(22, 6) * season_mult * year_mult)
        n_pedidos = max(4, n_pedidos)

        for _ in range(n_pedidos):
            # Vipar: 2% dos pedidos mas com valor 10-15x maior (alvo: ~19% da receita)
            if random.random() < 0.022:
                cliente = vipar
                vendedor = next(v for v in VENDEDORES if v["nome"] == "Carla Mendes")
            else:
                cliente = random.choice(ativos)
                vendedor_nome = cliente["vendedor_responsavel"]
                vendedor = next((v for v in VENDEDORES if v["nome"] == vendedor_nome), random.choice(VENDEDORES))

            # número de itens no pedido — médio de 3 itens
            n_itens = random.choices([1, 2, 3, 4, 5, 6], weights=[15, 28, 28, 15, 9, 5])[0]

            for _ in range(n_itens):
                produto = random.choice(produtos)
                qtd_base = {
                    "RL": (1, 4),
                    "UN": (1, 20),
                    "CX": (1, 2),
                }.get(produto["unidade"], (1, 5))

                # Vipar compra em volume alto
                if cliente["cliente_id"] == "AUR-0001":
                    qtd = random.randint(qtd_base[0] * 6, qtd_base[1] * 10)
                else:
                    qtd = random.randint(*qtd_base)

                # preço com variação
                preco_unit = produto["preco_tabela"] * random.uniform(0.92, 1.05)

                # desconto: Vipar tem desconto agressivo (4-10%); outros menos
                if cliente["cliente_id"] == "AUR-0001":
                    desconto = random.uniform(0.04, 0.10)
                elif cliente["segmento"] == "Construtora":
                    desconto = random.uniform(0.02, 0.06)
                else:
                    desconto = random.uniform(0.00, 0.04)

                valor_total = round_money(qtd * preco_unit * (1 - desconto))

                # margem: LED degrada ao longo do tempo (anomalia plantada)
                custo_base = produto["custo_medio"]
                if produto["categoria"] == "Iluminação LED":
                    # de jan-2024 (custo base) até dez-2025 → custo cresce ~+25%
                    months_elapsed = (d.year - 2024) * 12 + (d.month - 1)
                    custo_mult = 1 + (months_elapsed / 24) * 0.25
                    custo_base = custo_base * custo_mult
                # Prysmian +23% nos últimos 6 meses (jul/2025 em diante)
                if produto["fornecedor_principal"] == "Prysmian" and d >= date(2025, 7, 1):
                    custo_base = custo_base * 1.23

                custo_total = round_money(qtd * custo_base)

                # forma e prazo
                if cliente["cliente_id"] == "AUR-0001":
                    forma = "Boleto"
                    prazo = random.choice([45, 60, 60, 75, 90])  # prazos longos
                elif cliente["segmento"] == "Construtora":
                    forma = random.choice(["Boleto", "Boleto", "PIX"])
                    prazo = random.choice([30, 45, 45, 60])
                elif cliente["segmento"] == "Instaladora":
                    forma = random.choice(["PIX", "Boleto", "Cartão"])
                    prazo = random.choice([15, 28, 30])
                else:
                    forma = random.choice(["PIX", "Boleto", "Cartão"])
                    prazo = random.choice([0, 15, 28, 30])

                # Marcos Tavares — ticket médio menor (~40% abaixo): força fator
                if vendedor["nome"] == "Marcos Tavares":
                    valor_total = valor_total * 0.55
                    custo_total = custo_total * 0.55

                vendas.append({
                    "pedido_id": f"PED-{pedido_id_seq}",
                    "data_emissao": d.isoformat(),
                    "cliente_id": cliente["cliente_id"],
                    "razao_social": cliente["razao_social"],
                    "vendedor_id": vendedor["id"],
                    "vendedor": vendedor["nome"],
                    "regiao": vendedor["regiao"],
                    "sku": produto["sku"],
                    "descricao": produto["descricao"],
                    "categoria": produto["categoria"],
                    "fornecedor": produto["fornecedor_principal"],
                    "quantidade": qtd,
                    "preco_unit": round_money(preco_unit),
                    "desconto_pct": round(desconto * 100, 2),
                    "valor_total": round_money(valor_total),
                    "custo_total": round_money(custo_total),
                    "margem_bruta_pct": round((valor_total - custo_total) / valor_total * 100, 2) if valor_total > 0 else 0,
                    "forma_pagamento": forma,
                    "prazo_dias": prazo,
                    "status": "Faturado",
                })

            pedido_id_seq += 1

    # Cancelar/devolver ~3% dos pedidos
    for v in vendas:
        if random.random() < 0.02:
            v["status"] = "Cancelado"
        elif random.random() < 0.012:
            v["status"] = "Devolvido"

    return vendas

# ============== GERAR FINANCEIRO ==============

def gen_dre(vendas: list[dict]) -> list[dict]:
    """DRE mensal últimos 12 meses (jan-2025 → dez-2025)."""
    linhas = []
    meses = [date(2025, m, 1) for m in range(1, 13)]

    # Calcular receita real de vendas faturadas
    receita_por_mes = {}
    custo_por_mes = {}
    for v in vendas:
        if v["status"] != "Faturado":
            continue
        d = date.fromisoformat(v["data_emissao"])
        if d.year != 2025:
            continue
        mes = d.month
        receita_por_mes[mes] = receita_por_mes.get(mes, 0) + v["valor_total"]
        custo_por_mes[mes] = custo_por_mes.get(mes, 0) + v["custo_total"]

    for mes in meses:
        m = mes.month
        receita_bruta = receita_por_mes.get(m, 3800000)
        # impostos s/ venda ~12% (PIS, COFINS, ICMS médio)
        impostos = receita_bruta * 0.12
        receita_liquida = receita_bruta - impostos
        cmv = custo_por_mes.get(m, receita_liquida * 0.74)
        lucro_bruto = receita_liquida - cmv
        margem_bruta = lucro_bruto / receita_liquida if receita_liquida else 0

        # despesas operacionais — frete cresce ao longo do ano (anomalia)
        frete_pct = 0.021 + (m / 12) * 0.017  # 2.1% → 3.8%
        frete = receita_liquida * frete_pct
        comissoes = receita_bruta * 0.022
        salarios = 220000 + m * 2000  # leve crescimento
        aluguel = 78000
        marketing = 32000 + (15000 if m in (3, 9) else 0)  # picos
        energia = 18000 + (m * 400)
        outras_adm = 42000

        # despesas financeiras crescem 41% (capital de giro)
        desp_financeira = 12000 * (1 + (m - 1) / 11 * 0.41)
        depreciacao = 24500

        ebitda = lucro_bruto - (comissoes + salarios + aluguel + frete + marketing + energia + outras_adm)
        lucro_operacional = ebitda - depreciacao
        lucro_antes_ir = lucro_operacional - desp_financeira
        ir = max(0, lucro_antes_ir * 0.15)
        lucro_liquido = lucro_antes_ir - ir

        linhas.extend([
            {"mes_ref": mes.isoformat(), "linha_dre": "Receita Bruta",         "categoria": "Receita",       "valor": round_money(receita_bruta)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Impostos s/ Venda", "categoria": "Dedução",       "valor": round_money(-impostos)},
            {"mes_ref": mes.isoformat(), "linha_dre": "Receita Líquida",       "categoria": "Receita",       "valor": round_money(receita_liquida)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) CMV",               "categoria": "CMV",           "valor": round_money(-cmv)},
            {"mes_ref": mes.isoformat(), "linha_dre": "Lucro Bruto",           "categoria": "Resultado",     "valor": round_money(lucro_bruto)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Comissões",         "categoria": "Operacional",   "valor": round_money(-comissoes)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Salários+Encargos", "categoria": "Operacional",   "valor": round_money(-salarios)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Aluguel CD",        "categoria": "Operacional",   "valor": round_money(-aluguel)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Frete Saída",       "categoria": "Operacional",   "valor": round_money(-frete)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Marketing",         "categoria": "Operacional",   "valor": round_money(-marketing)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Energia",           "categoria": "Operacional",   "valor": round_money(-energia)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Outras Adm",        "categoria": "Operacional",   "valor": round_money(-outras_adm)},
            {"mes_ref": mes.isoformat(), "linha_dre": "EBITDA",                "categoria": "Resultado",     "valor": round_money(ebitda)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Depreciação",       "categoria": "Operacional",   "valor": round_money(-depreciacao)},
            {"mes_ref": mes.isoformat(), "linha_dre": "Lucro Operacional",     "categoria": "Resultado",     "valor": round_money(lucro_operacional)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) Desp. Financeiras", "categoria": "Financeira",    "valor": round_money(-desp_financeira)},
            {"mes_ref": mes.isoformat(), "linha_dre": "Lucro Antes IR",        "categoria": "Resultado",     "valor": round_money(lucro_antes_ir)},
            {"mes_ref": mes.isoformat(), "linha_dre": "(-) IR/CSLL",           "categoria": "Tributário",    "valor": round_money(-ir)},
            {"mes_ref": mes.isoformat(), "linha_dre": "Lucro Líquido",         "categoria": "Resultado",     "valor": round_money(lucro_liquido)},
        ])
    return linhas


def gen_fluxo_caixa() -> list[dict]:
    """Fluxo diário últimos 180 dias."""
    linhas = []
    end = date(2025, 12, 31)
    start = end - timedelta(days=180)
    saldo = 480000.0

    for d in days_iter(start, end):
        # entrada principal: recebimentos
        if d.weekday() < 5:
            # 3-6 entradas por dia útil
            n_in = random.randint(3, 7)
            for _ in range(n_in):
                v = round(random.gauss(28000, 12000), 2)
                if v < 500:
                    v = 500
                saldo += v
                linhas.append({
                    "data": d.isoformat(),
                    "tipo": "Entrada",
                    "categoria": random.choice(["Recebimento Cliente", "Recebimento Cliente", "Recebimento Cliente", "Outras Receitas"]),
                    "descricao": "Liquidação boleto" if random.random() > 0.4 else "PIX cliente",
                    "valor": v,
                    "saldo_acumulado": round(saldo, 2),
                    "conta": random.choice(["Itaú-PJ", "Itaú-PJ", "BB-PJ", "Caixa"]),
                })
        # saídas
        n_out = random.randint(2, 6)
        for _ in range(n_out):
            cat = random.choices(
                ["Pagamento Fornecedor", "Folha", "Aluguel", "Impostos", "Frete", "Energia/Util", "Marketing", "Outras"],
                weights=[40, 18, 6, 12, 8, 5, 6, 5],
            )[0]
            v = -abs(round(random.gauss(14000, 8000), 2))
            if cat == "Folha" and d.day in (5, 6):
                v = -abs(round(random.gauss(180000, 20000), 2))
            elif cat == "Aluguel" and d.day == 10:
                v = -78000
            saldo += v
            linhas.append({
                "data": d.isoformat(),
                "tipo": "Saída",
                "categoria": cat,
                "descricao": cat,
                "valor": v,
                "saldo_acumulado": round(saldo, 2),
                "conta": random.choice(["Itaú-PJ", "Itaú-PJ", "BB-PJ"]),
            })
    return linhas


def gen_contas_receber(vendas: list[dict]) -> list[dict]:
    """Contas a receber baseadas em vendas faturadas 2025."""
    titulos = []
    today = date(2025, 12, 31)
    titulo_seq = 50000

    for v in vendas:
        if v["status"] != "Faturado":
            continue
        emissao = date.fromisoformat(v["data_emissao"])
        if emissao.year != 2025:
            continue

        vencimento = emissao + timedelta(days=v["prazo_dias"])
        valor = v["valor_total"]

        # 60% pagos em dia, 25% pagos com atraso, 10% vencidos abertos, 5% renegociados
        r = random.random()
        if r < 0.55:
            # pago em dia ou até 5 dias depois
            data_pag = vencimento + timedelta(days=random.randint(-2, 5))
            if data_pag > today:
                data_pag = today
            status = "Pago"
            valor_pago = valor
            atraso = max(0, (data_pag - vencimento).days)
        elif r < 0.78:
            # atraso 6-45d
            atraso_d = random.randint(6, 45)
            data_pag = vencimento + timedelta(days=atraso_d)
            if data_pag > today:
                # ainda em aberto
                if vencimento < today:
                    status = "Vencido"
                    data_pag = None
                    valor_pago = 0
                    atraso = (today - vencimento).days
                else:
                    status = "A Vencer"
                    data_pag = None
                    valor_pago = 0
                    atraso = 0
            else:
                status = "Pago"
                valor_pago = valor
                atraso = atraso_d
        elif r < 0.92:
            # em aberto
            if vencimento < today:
                # piora ao longo do ano — vipar concentra atrasos
                if v["cliente_id"] == "AUR-0001":
                    # vipar com R$ ~480k vencidos
                    atraso = (today - vencimento).days
                else:
                    atraso = (today - vencimento).days
                status = "Vencido"
                valor_pago = 0
                data_pag = None
            else:
                status = "A Vencer"
                valor_pago = 0
                data_pag = None
                atraso = 0
        else:
            status = "Renegociado"
            valor_pago = valor * 0.5
            data_pag = None
            atraso = 0

        titulos.append({
            "titulo_id": f"REC-{titulo_seq}",
            "pedido_id": v["pedido_id"],
            "cliente_id": v["cliente_id"],
            "razao_social": v["razao_social"],
            "data_emissao": v["data_emissao"],
            "data_vencimento": vencimento.isoformat(),
            "data_pagamento": data_pag.isoformat() if data_pag else "",
            "valor_titulo": valor,
            "valor_pago": valor_pago,
            "status": status,
            "dias_atraso": atraso,
            "forma_pagamento": v["forma_pagamento"],
        })
        titulo_seq += 1
    return titulos


def gen_contas_pagar() -> list[dict]:
    titulos = []
    today = date(2025, 12, 31)
    titulo_seq = 70000

    fornecedores = [
        ("Prysmian Cabos do Brasil", "Insumos", 0.31),
        ("Schneider Electric Brasil", "Insumos", 0.14),
        ("Empalux Iluminação", "Insumos", 0.09),
        ("Avant Iluminação", "Insumos", 0.06),
        ("Tigre S.A.", "Insumos", 0.08),
        ("Intelbras", "Insumos", 0.05),
        ("Steck do Brasil", "Insumos", 0.07),
        ("WEG Equipamentos", "Insumos", 0.04),
        ("Folha de Pagamento", "Folha", 0.06),
        ("Auriverde Imóveis", "Aluguel", 0.02),
        ("CPFL Energia", "Energia", 0.01),
        ("Marketing Digital Adoro", "Marketing", 0.005),
        ("Receita Federal", "Impostos", 0.05),
        ("Sefaz SP", "Impostos", 0.025),
    ]

    for mes in range(1, 13):
        for forn, cat, peso in fornecedores:
            n_titulos = max(1, int(peso * 40))
            for _ in range(n_titulos):
                d = date(2025, mes, random.randint(1, 28))
                vencimento = d + timedelta(days=random.choice([15, 28, 30, 45]))
                valor = -abs(round(random.gauss(15000 * (peso * 10 + 1), 8000), 2))
                # Prysmian +23% nos últimos 6 meses
                if "Prysmian" in forn and mes >= 7:
                    valor = valor * 1.23

                if vencimento <= today:
                    status = "Pago" if random.random() < 0.92 else "Em atraso"
                else:
                    status = "A vencer"

                titulos.append({
                    "titulo_id": f"PAG-{titulo_seq}",
                    "fornecedor": forn,
                    "cnpj_fornecedor": gen_cnpj(titulo_seq % 999),
                    "categoria_despesa": cat,
                    "data_emissao": d.isoformat(),
                    "data_vencimento": vencimento.isoformat(),
                    "valor": round_money(valor),
                    "status": status,
                })
                titulo_seq += 1
    return titulos

# ============== ESCREVER XLSX ==============

def write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict], table_name: str) -> None:
    ws = wb.create_sheet(name)
    write_headers(ws, headers)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(h))
    autosize(ws)
    add_table(ws, len(headers), len(rows) + 1, table_name)


def write_readme(wb: Workbook, title: str, paragraphs: list[str], dores: list[str]) -> None:
    ws = wb.create_sheet("Leia-me", 0)
    ws.column_dimensions["A"].width = 110

    title_font = Font(bold=True, size=16, color="0F1828", name="Inter")
    label_font = Font(bold=True, size=11, color="3B6BD9", name="Inter")
    body_font = Font(size=10, color="2A3447", name="Inter")
    italic_font = Font(italic=True, size=10, color="6B7280", name="Inter")

    ws["A1"] = title
    ws["A1"].font = title_font

    ws["A3"] = "Empresa fictícia · dados gerados para fins didáticos"
    ws["A3"].font = italic_font

    ws["A5"] = "Sobre a empresa"
    ws["A5"].font = label_font

    row = 6
    for p in paragraphs:
        ws.cell(row=row, column=1, value=p).font = body_font
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Padrões plantados — descubra com Claude").font = label_font
    row += 1
    for d in dores:
        c = ws.cell(row=row, column=1, value=f"• {d}")
        c.font = body_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Gerado pela openacademy · openacademyai.jvictordev.expert").font = italic_font


def build_vendas_xlsx(clientes, produtos, vendas) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    write_readme(
        wb,
        "Distribuidora Aurora — Vendas 2024-2025",
        [
            "Distribuidora B2B de materiais elétricos para construtoras, instaladoras e revendas.",
            f"{COMPANY['endereco']} · Filial: {COMPANY['filial']} · 87 funcionários · Faturamento ~R$ 3,8M/mês",
            "11 vendedores (8 externos + 3 inside sales) · ~380 clientes ativos · 4 categorias de produto.",
            "Dados de 2024-2025: dezenas de milhares de pedidos, todos os clientes, todos os SKUs.",
        ],
        [
            "Concentração: 1 cliente top contribui com fatia desproporcional da receita (e prazo de pagamento longo).",
            "Vendedor com performance fora da curva — ticket médio muito abaixo do time, queda mês a mês.",
            "Margem de uma categoria caindo ao longo de 24 meses — efeito câmbio + concorrência.",
            "Sazonalidade clara em meses específicos (obras pré-chuva, festas).",
            "Cohort de retenção: clientes adquiridos em períodos diferentes têm comportamentos distintos.",
        ],
    )

    write_sheet(wb, "Vendedores",
                ["id", "nome", "regiao", "tipo", "data_admissao", "meta_mensal", "comissao_pct"],
                [{"id": v["id"], "nome": v["nome"], "regiao": v["regiao"], "tipo": v["tipo"], "data_admissao": v["data_admissao"], "meta_mensal": v["meta_mensal"], "comissao_pct": v["comissao_pct"]} for v in VENDEDORES],
                "TblVendedores")

    write_sheet(wb, "Clientes",
                ["cliente_id", "razao_social", "cnpj", "segmento", "cidade", "uf", "vendedor_responsavel", "data_primeira_compra", "limite_credito", "status"],
                clientes,
                "TblClientes")

    write_sheet(wb, "Produtos",
                ["sku", "descricao", "categoria", "subcategoria", "unidade", "custo_medio", "preco_tabela", "margem_alvo_pct", "fornecedor_principal"],
                produtos,
                "TblProdutos")

    write_sheet(wb, "Vendas",
                ["pedido_id", "data_emissao", "cliente_id", "razao_social", "vendedor_id", "vendedor", "regiao", "sku", "descricao", "categoria", "fornecedor", "quantidade", "preco_unit", "desconto_pct", "valor_total", "custo_total", "margem_bruta_pct", "forma_pagamento", "prazo_dias", "status"],
                vendas,
                "TblVendas")

    wb.save(OUT_DIR / "vendas-aurora-2024-2025.xlsx")


def build_financeiro_xlsx(dre, fluxo, receber, pagar) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    write_readme(
        wb,
        "Distribuidora Aurora — Financeiro 2025",
        [
            "Visão financeira anual da Aurora: DRE mensal, fluxo de caixa diário (180 dias), contas a receber e pagar.",
            "Aurora opera com capital de giro relevante — clientes pagam em 30-90 dias, fornecedores pagam em 28-45 dias.",
            "Concentração tributária Simples Nacional + ICMS. Aluguel CD em Cotia + filial em Campinas.",
        ],
        [
            "Frete saída crescendo silenciosamente ao longo do ano — vira ralo de margem.",
            "Despesas financeiras crescem com bancos cobrando capital de giro caro.",
            "Custo de fornecedor principal subiu nos últimos meses, impactando CMV.",
            "Atraso médio de recebimento crescendo — cliente top concentra valor vencido.",
            "Margem bruta comprimida — efeito CMV + frete + desconto.",
        ],
    )

    write_sheet(wb, "DRE_Mensal",
                ["mes_ref", "linha_dre", "categoria", "valor"],
                dre,
                "TblDRE")

    write_sheet(wb, "Fluxo_Caixa",
                ["data", "tipo", "categoria", "descricao", "valor", "saldo_acumulado", "conta"],
                fluxo,
                "TblFluxo")

    write_sheet(wb, "Contas_Receber",
                ["titulo_id", "pedido_id", "cliente_id", "razao_social", "data_emissao", "data_vencimento", "data_pagamento", "valor_titulo", "valor_pago", "status", "dias_atraso", "forma_pagamento"],
                receber,
                "TblReceber")

    write_sheet(wb, "Contas_Pagar",
                ["titulo_id", "fornecedor", "cnpj_fornecedor", "categoria_despesa", "data_emissao", "data_vencimento", "valor", "status"],
                pagar,
                "TblPagar")

    wb.save(OUT_DIR / "financeiro-aurora-2025.xlsx")

# ============== ALSO CSV ==============

def write_csv(name: str, headers: list[str], rows: list[dict]) -> None:
    import csv
    path = OUT_DIR / "csv" / f"{name}.csv"
    path.parent.mkdir(exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for r in rows:
            writer.writerow({h: r.get(h, "") for h in headers})

# ============== MAIN ==============

def main() -> None:
    print("Generating Aurora data...")
    clientes = gen_clientes()
    print(f"  Clientes: {len(clientes)}")
    produtos = gen_produtos()
    print(f"  Produtos: {len(produtos)}")
    vendas = gen_vendas(clientes, produtos)
    print(f"  Vendas: {len(vendas)}")
    dre = gen_dre(vendas)
    print(f"  DRE rows: {len(dre)}")
    fluxo = gen_fluxo_caixa()
    print(f"  Fluxo: {len(fluxo)}")
    receber = gen_contas_receber(vendas)
    print(f"  Contas a Receber: {len(receber)}")
    pagar = gen_contas_pagar()
    print(f"  Contas a Pagar: {len(pagar)}")

    print("Writing xlsx files...")
    build_vendas_xlsx(clientes, produtos, vendas)
    build_financeiro_xlsx(dre, fluxo, receber, pagar)

    print("Writing CSV mirrors...")
    write_csv("vendedores",        ["id", "nome", "regiao", "tipo", "data_admissao", "meta_mensal", "comissao_pct"], VENDEDORES)
    write_csv("clientes",          ["cliente_id", "razao_social", "cnpj", "segmento", "cidade", "uf", "vendedor_responsavel", "data_primeira_compra", "limite_credito", "status"], clientes)
    write_csv("produtos",          ["sku", "descricao", "categoria", "subcategoria", "unidade", "custo_medio", "preco_tabela", "margem_alvo_pct", "fornecedor_principal"], produtos)
    write_csv("vendas",            ["pedido_id", "data_emissao", "cliente_id", "razao_social", "vendedor_id", "vendedor", "regiao", "sku", "descricao", "categoria", "fornecedor", "quantidade", "preco_unit", "desconto_pct", "valor_total", "custo_total", "margem_bruta_pct", "forma_pagamento", "prazo_dias", "status"], vendas)
    write_csv("dre",               ["mes_ref", "linha_dre", "categoria", "valor"], dre)
    write_csv("fluxo_caixa",       ["data", "tipo", "categoria", "descricao", "valor", "saldo_acumulado", "conta"], fluxo)
    write_csv("contas_receber",    ["titulo_id", "pedido_id", "cliente_id", "razao_social", "data_emissao", "data_vencimento", "data_pagamento", "valor_titulo", "valor_pago", "status", "dias_atraso", "forma_pagamento"], receber)
    write_csv("contas_pagar",      ["titulo_id", "fornecedor", "cnpj_fornecedor", "categoria_despesa", "data_emissao", "data_vencimento", "valor", "status"], pagar)

    # sanity check
    from collections import defaultdict
    total_2025 = sum(v["valor_total"] for v in vendas if v["status"] == "Faturado" and v["data_emissao"].startswith("2025"))
    vipar_2025 = sum(v["valor_total"] for v in vendas if v["status"] == "Faturado" and v["data_emissao"].startswith("2025") and v["cliente_id"] == "AUR-0001")

    # ticket médio por pedido por vendedor
    pedido_valor = defaultdict(float)
    pedido_vendedor = {}
    for v in vendas:
        if v["status"] != "Faturado":
            continue
        pedido_valor[v["pedido_id"]] += v["valor_total"]
        pedido_vendedor[v["pedido_id"]] = v["vendedor"]

    vendedor_pedidos = defaultdict(list)
    for pid, val in pedido_valor.items():
        vendedor_pedidos[pedido_vendedor[pid]].append(val)

    print()
    print("===== SANITY CHECK =====")
    print(f"Total pedidos:  {len(pedido_valor):,}")
    print(f"Total linhas:   {len(vendas):,}")
    print(f"Receita 2025:   R$ {total_2025:,.0f}")
    print(f"Vipar 2025:     R$ {vipar_2025:,.0f}  ({vipar_2025 / total_2025 * 100:.1f}%)")
    print()
    print("Ticket médio por pedido por vendedor:")
    for v in VENDEDORES:
        pedidos = vendedor_pedidos.get(v["nome"], [])
        if pedidos:
            tm = sum(pedidos) / len(pedidos)
            print(f"  {v['nome']:25} R$ {tm:>8,.0f} ({len(pedidos):>4} pedidos)")
    print()
    print(f"Files in: {OUT_DIR}")


if __name__ == "__main__":
    main()
